import openpyxl as xl
import shutil
import argparse
from tqdm import tqdm

parser = argparse.ArgumentParser()

parser.add_argument('--filename', required=True)
parser.add_argument('--output_dir', default='static/xlsx')

args = parser.parse_args()

FRI_SUN_FILE_NAME = "fri_sunny_shift.xlsx"
FRI_RAIN_FILE_NAME = "fri_rain_shift.xlsx"
SAT_SUN_FILE_NAME = "sat_sunny_shift.xlsx"
SAT_RAIN_FILE_NAME = "sat_rain_shift.xlsx"
SUN_SUN_FILE_NAME = "sun_sunny_shift.xlsx"
SUN_RAIN_FILE_NAME = "sun_rain_shift.xlsx"
MON_SUN_FILE_NAME = "mon_sunny_shift.xlsx"
MON_RAIN_FILE_NAME = "mon_rain_shift.xlsx"
COPY_FILE_NAME = "copy.xlsx"

TUE_SHEET_NAME = "準々備日"
FRI_SUN_SHEET_NAME = "準備日晴れ"
FRI_RAIN_SHEET_NAME = "準備日雨"
SAT_SUN_SHEET_NAME = "1日目晴れ"
SAT_RAIN_SHEET_NAME = "1日目雨"
SUN_SUN_SHEET_NAME = "2日目晴れ"
SUN_RAIN_SHEET_NAME = "2日目雨"
MON_SUN_SHEET_NAME = "片付け日"
MON_RAIN_SHEET_NAME = "片付け日"

DELETE_SHEET_NAME = [
    '準々備日',
    '準備日晴れ',
    '準備日雨',
    '1日目晴れ',
    '1日目雨',
    '2日目晴れ',
    '2日目雨',
    '片付け日',
    '片付け日'
]


def delete_and_output(wb, sheet, file):
    for j in range(len(wb.sheetnames)):
        if DELETE_SHEET_NAME[j] != sheet:
            wb.remove(wb[DELETE_SHEET_NAME[j]])
    wb.save(args.output_dir + '/' + file)


book = xl.load_workbook(args.filename)

for i in tqdm(range(len(book.sheetnames))):
    shutil.copy(args.filename, args.output_dir + "/" + COPY_FILE_NAME)
    wb = xl.load_workbook(args.output_dir + '/' + COPY_FILE_NAME)

    if book.worksheets[i].title == FRI_SUN_SHEET_NAME:
        delete_and_output(wb, FRI_SUN_SHEET_NAME, FRI_SUN_FILE_NAME)
        continue
    elif book.worksheets[i].title == FRI_RAIN_SHEET_NAME:
        delete_and_output(wb, FRI_RAIN_SHEET_NAME, FRI_RAIN_FILE_NAME)
        continue
    elif book.worksheets[i].title == SAT_SUN_SHEET_NAME:
        delete_and_output(wb, SAT_SUN_SHEET_NAME, SAT_SUN_FILE_NAME)
        continue
    elif book.worksheets[i].title == SAT_RAIN_SHEET_NAME:
        delete_and_output(wb, SAT_RAIN_SHEET_NAME, SAT_RAIN_FILE_NAME)
        continue
    elif book.worksheets[i].title == SUN_SUN_SHEET_NAME:
        delete_and_output(wb, SUN_SUN_SHEET_NAME, SUN_SUN_FILE_NAME)
        continue
    elif book.worksheets[i].title == SUN_RAIN_SHEET_NAME:
        delete_and_output(wb, SUN_RAIN_SHEET_NAME, SUN_RAIN_FILE_NAME)
        continue
    elif book.worksheets[i].title == MON_SUN_SHEET_NAME:
        delete_and_output(wb, MON_SUN_SHEET_NAME, MON_SUN_FILE_NAME)
        ws = wb.active
        ws.title = '片付け日晴れ'
        wb.save(args.output_dir + '/' + MON_SUN_FILE_NAME)
        continue
    elif book[MON_SUN_SHEET_NAME].title == MON_RAIN_SHEET_NAME:
        delete_and_output(wb, MON_RAIN_SHEET_NAME, MON_RAIN_FILE_NAME)
        ws = wb.active
        ws.title = '片付け日雨'
        wb.save(args.output_dir + '/' + MON_RAIN_FILE_NAME)
